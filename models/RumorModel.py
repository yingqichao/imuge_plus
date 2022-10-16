import logging
import torchvision
import torch.nn as nn
from torch.nn.parallel import DistributedDataParallel
from skimage.color import rgb2gray
import models.lr_scheduler as lr_scheduler
from .base_model import BaseModel
from models.modules.loss import ReconstructionLoss, CWLoss
from models.modules.Quantization import Quantization
import torch.distributed as dist
from losses.loss import AdversarialLoss, PerceptualLoss, StyleLoss
import cv2
from utils.metrics import PSNR
import os
import pytorch_ssim
from noise_layers import *
# import matlab.engine
from losses.loss import ExclusionLoss


logger = logging.getLogger('base')


class RumorModel(BaseModel):
    def __init__(self, opt):
        super(RumorModel, self).__init__(opt)

        self.IMG_EXTENSIONS = ['.jpg', '.JPG', '.jpeg', '.JPEG', '.png', '.PNG', '.ppm', '.PPM', '.bmp', '.BMP']
        ########### CONSTANTS ###############
        self.TASK_IMUGEV2 = "ImugeV2"
        self.TASK_TEST = "Test"
        self.TASK_CropLocalize = "CropLocalize"
        self.TASK_RHI3 = "RHI3"
        if opt['dist']:
            self.rank = torch.distributed.get_rank()
        else:
            self.rank = -1  # non dist training
        train_opt = opt['train']
        test_opt = opt['test']
        self.opt = opt
        # self.lpips_vgg = lpips.LPIPS(net="vgg").cuda()
        self.train_opt = train_opt
        self.test_opt = test_opt
        self.real_H, self.real_H_path, self.previous_images, self.previous_previous_images = None, None, None, None
        self.task_name = self.train_opt['task_name']
        print("Task Name: {}".format(self.task_name))
        self.global_step = 0
        self.new_task = self.train_opt['new_task']

        ############## Metrics and attacks #############
        self.tanh = nn.Tanh().cuda()
        self.psnr = PSNR(255.0).cuda()
        self.exclusion_loss = ExclusionLoss().type(torch.cuda.FloatTensor).cuda()
        self.ssim_loss = pytorch_ssim.SSIM().cuda()

        self.bce_loss = nn.BCELoss().cuda()
        self.l1_loss = nn.SmoothL1Loss().cuda()  # reduction="sum"
        self.l2_loss = nn.MSELoss().cuda()  # reduction="sum"
        self.perceptual_loss = PerceptualLoss().cuda()
        self.style_loss = StyleLoss().cuda()
        self.Quantization = Quantization().cuda()
        self.Reconstruction_forw = ReconstructionLoss(losstype=self.train_opt['pixel_criterion_forw']).cuda()
        self.Reconstruction_back = ReconstructionLoss(losstype=self.train_opt['pixel_criterion_back']).cuda()
        self.criterion_adv = CWLoss().cuda()  # loss for fooling target model
        self.criterion = nn.CrossEntropyLoss().cuda()
        self.width_height = opt['datasets']['train']['GT_size']
        self.init_gaussian = None
        self.adversarial_loss = AdversarialLoss(type="nsgan").cuda()

        ############## Nets ################################
        self.netG = torchvision.models.inception_v3(pretrained=True)
        self.netG.fc = nn.Sequential(
            nn.Linear(2048, 256),
            nn.ReLU(inplace=True),
            nn.Dropout(p=0.5, inplace=False),
            nn.Linear(256, 256),
            nn.ReLU(inplace=True),
            nn.Dropout(p=0.5, inplace=False),
            nn.Linear(256, 1),
        )
        self.netG = self.netG.cuda()
        self.netG = DistributedDataParallel(self.netG, device_ids=[torch.cuda.current_device()],
                                                 find_unused_parameters=True)

        ########### For Crop localization ############
        wd_G = train_opt['weight_decay_G'] if train_opt['weight_decay_G'] else 0

        ########## Load pre-trained ##################
        self.out_space_storage = '/home/qcying/20220420_Qian_rumor'
        self.model_path = '/model/Qian_rumor_dataset/3010'

        load_state = True
        if load_state:
            # 45010
            pretrain = self.out_space_storage + self.model_path

            load_path_G = pretrain + "_netG.pth"
            if load_path_G is not None:
                logger.info('Loading model for class [{:s}] ...'.format(load_path_G))
                if os.path.exists(load_path_G):
                    self.load_network(load_path_G, self.netG, self.opt['path']['strict_load'])
                else:
                    logger.info('Did not find model for class [{:s}] ...'.format(load_path_G))

        ########## optimizers ##################
        wd_G = train_opt['weight_decay_G'] if train_opt['weight_decay_G'] else 0

        optim_params = []
        for k, v in self.netG.named_parameters():
            if v.requires_grad:
                optim_params.append(v)
            else:
                if self.rank <= 0:
                    logger.warning('Params [{:s}] will not optimize.'.format(k))
        self.optimizer_G = torch.optim.AdamW(optim_params, lr=train_opt['lr_G'],
                                             weight_decay=wd_G,
                                             betas=(train_opt['beta1'], train_opt['beta2']))
        self.optimizers.append(self.optimizer_G)

        # ############## schedulers #########################
        if train_opt['lr_scheme'] == 'MultiStepLR':
            for optimizer in self.optimizers:
                self.schedulers.append(
                    lr_scheduler.MultiStepLR_Restart(optimizer, train_opt['lr_steps'],
                                                     restarts=train_opt['restarts'],
                                                     weights=train_opt['restart_weights'],
                                                     gamma=train_opt['lr_gamma'],
                                                     clear_state=train_opt['clear_state']))
        elif train_opt['lr_scheme'] == 'CosineAnnealingLR_Restart':
            for optimizer in self.optimizers:
                self.schedulers.append(
                    lr_scheduler.CosineAnnealingLR_Restart(
                        optimizer, train_opt['T_period'], eta_min=train_opt['eta_min'],
                        restarts=train_opt['restarts'], weights=train_opt['restart_weights']))
        else:
            raise NotImplementedError('MultiStepLR learning rate scheme is enough.')

        ######## init constants
        self.forward_image_buff = None

    def feed_data(self, batch):

        img, label = batch
        self.real_H = img.cuda()
        self.real_H = torch.clamp(self.real_H, 0, 1)
        self.label = label.cuda()

    def symm_pad(self, im, padding):
        h, w = im.shape[-2:]
        left, right, top, bottom = padding

        x_idx = np.arange(-left, w + right)
        y_idx = np.arange(-top, h + bottom)

        x_pad = self.reflect(x_idx, -0.5, w - 0.5)
        y_pad = self.reflect(y_idx, -0.5, h - 0.5)
        xx, yy = np.meshgrid(x_pad, y_pad)
        return im[..., yy, xx]

    def reflect(self, x, minx, maxx):
        """ Reflects an array around two points making a triangular waveform that ramps up
        and down,  allowing for pad lengths greater than the input length """
        rng = maxx - minx
        double_rng = 2 * rng
        mod = np.fmod(x - minx, double_rng)
        normed_mod = np.where(mod < 0, mod + double_rng, mod)
        out = np.where(normed_mod >= rng, double_rng - normed_mod, normed_mod) + minx
        return np.array(out, dtype=x.dtype)

    def optimize_parameters(self, step, latest_values=None, train=True, eval_dir=None):
        self.netG.train()
        self.global_step = self.global_step + 1
        logs, debug_logs = [], []

        self.real_H = torch.clamp(self.real_H, 0, 1)
        batch_size = self.real_H.shape[0]

        save_interval = 1000

        with torch.enable_grad():
            self.label = self.label.unsqueeze(1)
            predict_result, _ = self.netG(self.real_H) # Inception returns a hidden state
            # predict_result = torch.clamp(predict_result, 0, 1)


            lr = self.get_current_learning_rate()
            logs.append(('lr', lr))


            loss = self.l2_loss(predict_result, self.label)
            self.netG.train()
            self.optimizer_G.zero_grad()
            loss.backward()
            # if self.train_opt['gradient_clipping']:
            #     nn.utils.clip_grad_norm_(self.localizer.parameters(), self.train_opt['gradient_clipping'])
            self.optimizer_G.step()
            logs.append(('loss', loss.item()))



        ######## Finally ####################
        if step % save_interval == 10 and self.rank <= 0:
            logger.info('Saving models and training states.')
            self.save(self.global_step)

        return logs, debug_logs

    def is_image_file(self, filename):
        return any(filename.endswith(extension) for extension in self.IMG_EXTENSIONS)

    def get_paths_from_images(self, path):
        '''get image path list from image folder'''
        assert os.path.isdir(path), '{:s} is not a valid directory'.format(path)
        images = []
        for dirpath, _, fnames in sorted(os.walk(path)):
            for fname in sorted(fnames):
                if self.is_image_file(fname):
                    # img_path = os.path.join(dirpath, fname)
                    images.append((path, dirpath[len(path) + 1:], fname))
        assert images, '{:s} has no valid image file'.format(path)
        return images

    def evaluate(self):
        import os
        import openpyxl
        from tqdm import tqdm
        self.paths_GT, self.label_dict = [], {}
        wb = openpyxl.load_workbook('/home/groupshare/Distinguish_rumors_using_attached_images/labels_image.xlsx')
        folder = '/home/groupshare/rumor_dataset/rumor_datasets/images/All images/'
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        self.num_consist, self.num_correct = 0., 0.
        self.netG.eval()
        dir_name = '/home/groupshare/COVID_images/real_sum/'
        images = os.listdir(dir_name)
        num_valid = len(images)
        for i in tqdm(range(len(images))):
            # image_name = sheet['B' + str(i)].value
            try:
                img_GT = self.load_image(dir_name + images[i], grayscale=False)
                img_GT = self.img_random_crop(img_GT, 512, 512).cuda().unsqueeze(0)
                img_GT = torch.clamp(img_GT, 0, 1)
                output  = self.netG(img_GT)
                output = output.item()
                predict = round(output)
                ground_truth = 1 #int(sheet['C' + str(i)].value)
                # human_label = round(float(sheet['E' + str(i)].value))
                if predict==ground_truth: self.num_correct = self.num_correct+1
                print("{} {:4f}".format(images[i],output))
                # if predict == human_label: self.num_consist = self.num_consist + 1
            except Exception:
                print("Cannot open {}".format('./COVID_images/' + images[i]))
                num_valid -= 1
        print("Correct {} Sum {}".format(self.num_correct, num_valid))

        # print("Correct {} Consistent {} Sum {}".format(self.num_correct,self.num_consist,rows-1))

    def img_random_crop(self, img_GT, Height=512, Width=512, grayscale=False):
        # # randomly crop
        # H, W = img_GT.shape[0], img_GT.shape[1]
        # rnd_h = random.randint(0, max(0, H - Height))
        # rnd_w = random.randint(0, max(0, W - Width))
        #
        # img_GT = img_GT[rnd_h:rnd_h + Height, rnd_w:rnd_w + Width, :]
        #
        # orig_height, orig_width, _ = img_GT.shape
        # H, W = img_GT.shape[0], img_GT.shape[1]

        # BGR to RGB, HWC to CHW, numpy to tensor
        if not grayscale:
            img_GT = img_GT[:, :, [2, 1, 0]]
            img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        else:
            img_GT = self.image_to_tensor(img_GT)

        return img_GT.cuda()

    def load_image(self, path, readimg=False, Height=512, Width=512,grayscale=False):
        import data.util as util
        GT_path = path

        img_GT = util.read_img(GT_path)

        # change color space if necessary
        # img_GT = util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]
        if grayscale:
            img_GT = rgb2gray(img_GT)

        img_GT = cv2.resize(np.copy(img_GT), (Width, Height), interpolation=cv2.INTER_LINEAR)
        return img_GT

    def save(self, iter_label):
        self.save_network(self.netG, 'netG', iter_label,
                          model_path=self.out_space_storage + '/model/' + self.opt['datasets']['train']['name'] + '/')

