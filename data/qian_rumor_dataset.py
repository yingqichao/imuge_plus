import random
import numpy as np
import cv2
import lmdb
import torch
import torch.utils.data as data
import data.util as util
import os
# from turbojpeg import TurboJPEG
from PIL import Image
from jpeg2dct.numpy import load, loads
from skimage.feature import canny
from skimage.color import rgb2gray, gray2rgb
import torchvision.transforms.functional as F

import time
from PIL import Image, ImageTk
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from scp import SCPClient
from tqdm import tqdm
import datetime
import random
import shutil

class QianDataset(data.Dataset):
    '''
    Read LQ (Low Quality, here is LR) and GT image pairs.
    If only GT image is provided, generate LQ image on-the-fly.
    The pair is ensured by 'sorted' function, so please check the name convention.
    '''

    def __init__(self, opt, dataset_opt):
        super(QianDataset, self).__init__()
        self.opt = opt
        self.dataset_opt = dataset_opt
        self.paths_GT = None
        self.sizes_GT = None
        self.index = 0

        # self.paths_GT, self.sizes_GT = util.get_image_paths(dataset_opt['dataroot_GT'])
        self.paths_GT, self.label_dict = [], {}
        wb = openpyxl.load_workbook('/home/groupshare/Distinguish_rumors_using_attached_images/labels_image.xlsx')
        folder = '/home/groupshare/rumor_dataset/rumor_datasets/images/All images/'
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            image_name = sheet['B' + str(i)].value
            image_label = float(sheet['E' + str(i)].value)
            self.paths_GT.append(folder+image_name)
            self.label_dict[folder+image_name] = image_label

        assert self.paths_GT, 'Error: GT path is empty.'



    def __getitem__(self, index):
        # index = self.index
        # self.index +=1
        GT_size = self.dataset_opt['GT_size']

        # get GT image
        GT_path = self.paths_GT[index]
        label = self.label_dict[GT_path]

        try:
            img_GT = util.read_img(GT_path)
        except Exception:
            raise IOError("Load {} Error".format(GT_path))

        img_GT = util.channel_convert(img_GT.shape[2], self.dataset_opt['color'], [img_GT])[0]

        ###### directly resize instead of crop
        img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
                            interpolation=cv2.INTER_LINEAR)

        orig_height, orig_width, _ = img_GT.shape
        H, W, _ = img_GT.shape


        # BGR to RGB, HWC to CHW, numpy to tensor
        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]



        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        # print("Query ok {}".format(GT_path))

        return (img_GT, torch.tensor(label,dtype=torch.float32))

    def __len__(self):
        return len(self.paths_GT)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t
